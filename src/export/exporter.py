"""
CSV export utilities for compression metrics and comparison reports.
"""

import csv
from datetime import datetime
from pathlib import Path

import matplotlib
matplotlib.use("Agg")  # Non-interactive backend to avoid GUI freezes
import matplotlib.pyplot as plt


CSV_COLUMNS = [
    "file",
    "algorithm",
    "original_size_kb",
    "compressed_size_kb",
    "reduction_percent",
    "time_ms",
    "resolution",
    "status",
]

COMPARISON_CSV_COLUMNS = [
    "algorithm",
    "files_processed",
    "avg_reduction",
    "avg_time_ms",
    "success",
    "failed",
    "score",
]


def export_metrics_csv(metrics, summary, output_dir):
    """Export metrics to results_export.csv with the required columns."""
    reports_dir = Path(output_dir)
    reports_dir.mkdir(parents=True, exist_ok=True)
    output_path = reports_dir / "results_export.csv"

    # Pre-calculate best algorithm and scores for each file
    file_groups = {}
    for m in metrics:
        if m.get("status") == "SUCCESS":
            fname = m["file"]
            if fname not in file_groups:
                file_groups[fname] = []
            file_groups[fname].append(m)

    file_stats = {}
    for fname, group in file_groups.items():
        max_red = max(m["reduction_percent"] for m in group)
        min_time = min(m["time_ms"] for m in group)
        
        # Best algorithm has max reduction. Break ties with faster time.
        best_candidates = [m for m in group if m["reduction_percent"] == max_red]
        best_metric = min(best_candidates, key=lambda m: m["time_ms"])
        best_algo = best_metric["algorithm"]
        
        file_stats[fname] = {
            "max_reduction": max_red,
            "min_time": min_time,
            "best_algorithm": best_algo
        }

    headers = [
        "File Name",
        "Original Size",
        "Algorithm",
        "Compressed Size",
        "Reduction %",
        "Compression Ratio",
        "Processing Time (ms)",
        "Best Algorithm",
        "Score",
        "Timestamp"
    ]

    timestamp_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    with output_path.open("w", newline="", encoding="utf-8") as csv_file:
        writer = csv.writer(csv_file)
        writer.writerow(headers)
        
        for m in metrics:
            fname = m["file"]
            orig_sz = m["original_size"]
            comp_sz = m["compressed_size"]
            reduction = m["reduction_percent"]
            time_ms = m["time_ms"]
            algo = m["algorithm"]
            
            # Compression ratio: original_size / compressed_size
            ratio = orig_sz / comp_sz if comp_sz > 0 else 1.0
            
            # Find best algorithm and score
            if fname in file_stats and m.get("status") == "SUCCESS":
                stats = file_stats[fname]
                best_algo = stats["best_algorithm"]
                
                max_red = stats["max_reduction"]
                min_time = stats["min_time"]
                
                # Formula: 50% reduction + 50% speed (normalized)
                red_score = (reduction / max_red) * 50 if max_red > 0 else 0
                speed_score = (min_time / max(time_ms, 0.001)) * 50 if time_ms > 0 else 25
                score = min(red_score + speed_score, 100)
            else:
                best_algo = "-"
                score = 0.0
                
            writer.writerow([
                fname,
                orig_sz,
                algo,
                comp_sz,
                f"{reduction:.2f}%",
                f"{ratio:.2f}x",
                f"{time_ms:.2f}",
                best_algo,
                f"{score:.2f}",
                timestamp_str
            ])
            
    return output_path


def _timestamp_ms():
    """Return a millisecond-precision timestamp string: YYYYMMDD_HHMMSS_mmm."""
    now = datetime.now()
    return f"{now.strftime('%Y%m%d_%H%M%S')}_{now.microsecond // 1000:03d}"


def _build_report_path(reports_dir):
    """Build a timestamped report path with millisecond precision."""
    timestamp = _timestamp_ms()
    output_path = reports_dir / f"report_{timestamp}.csv"
    if not output_path.exists():
        return output_path

    counter = 1
    while True:
        candidate = reports_dir / f"report_{timestamp}_{counter}.csv"
        if not candidate.exists():
            return candidate
        counter += 1


def _format_metric_row(metric):
    """Convert one metrics object to a CSV row."""
    return {
        "file": metric["file"],
        "algorithm": metric["algorithm"],
        "original_size_kb": _bytes_to_kb(metric["original_size"]),
        "compressed_size_kb": _bytes_to_kb(metric["compressed_size"]),
        "reduction_percent": f"{metric['reduction_percent']:.4f}",
        "time_ms": f"{metric['time_ms']:.4f}",
        "resolution": metric["resolution"],
        "status": metric["status"],
    }


def _bytes_to_kb(value):
    """Convert bytes to kilobytes."""
    if value <= 0:
        return "0.0000"
    return f"{value / 1024:.4f}"


# -------------------------------------------------------------------------
# Comparison report export
# -------------------------------------------------------------------------


def export_comparison_csv(per_algorithm, winners, output_dir):
    """Export comparison results to a timestamped CSV file with scoring."""
    reports_dir = Path(output_dir)
    reports_dir.mkdir(parents=True, exist_ok=True)

    output_path = reports_dir / f"comparison_report_{_timestamp_ms()}.csv"

    # Compute scores for each algorithm
    scores = compute_ranking_scores(per_algorithm)

    with output_path.open("w", newline="", encoding="utf-8") as csv_file:
        writer = csv.writer(csv_file)

        # Header
        writer.writerow(COMPARISON_CSV_COLUMNS)

        # Per-algorithm rows (sorted by score descending)
        sorted_algos = sorted(
            per_algorithm.items(),
            key=lambda item: scores.get(item[0], 0),
            reverse=True,
        )
        for algo_key, data in sorted_algos:
            s = data["summary"]
            successful = sum(1 for m in data["metrics"] if m["status"] == "SUCCESS")
            failed_count = sum(1 for m in data["metrics"] if m["status"] == "FAILED")
            writer.writerow([
                data["label"],
                s["completed"],
                f"{s['avg_reduction']:.4f}",
                f"{s['avg_time_ms']:.4f}",
                successful,
                failed_count,
                f"{scores.get(algo_key, 0):.2f}",
            ])

        # Winners section
        writer.writerow([])
        writer.writerow(["=== WINNERS ==="])
        writer.writerow(["best_compression", winners.get("winner_reduction_label", "N/A")])
        writer.writerow(["best_compression_value", f"{winners.get('winner_reduction_value', 0):.4f}"])
        writer.writerow(["fastest", winners.get("winner_speed_label", "N/A")])
        writer.writerow(["fastest_value", f"{winners.get('winner_speed_value', 0):.4f}"])
        writer.writerow(["balanced", winners.get("winner_balanced_label", "N/A")])
        writer.writerow(["balanced_score", f"{winners.get('winner_balanced_value', 0):.2f}"])

    return output_path


def compute_ranking_scores(per_algorithm):
    """Compute a balanced ranking score (0-100) for each algorithm.

    Higher score = better overall performance.
    Formula: 50% reduction + 50% speed (normalized).
    """
    scores = {}
    reduction_values = {}
    time_values = {}

    for algo_key, data in per_algorithm.items():
        s = data["summary"]
        successful = [m for m in data["metrics"] if m["status"] == "SUCCESS"]
        if successful:
            reduction_values[algo_key] = s["avg_reduction"]
            time_values[algo_key] = s["avg_time_ms"]

    if not reduction_values:
        return scores

    max_reduction = max(reduction_values.values()) if reduction_values else 1
    min_time = min(time_values.values()) if time_values else 1

    for algo_key in reduction_values:
        red = reduction_values[algo_key]
        t = max(time_values[algo_key], 0.001)

        reduction_score = (red / max_reduction) * 50 if max_reduction > 0 else 0
        speed_score = (min_time / t) * 50 if min_time > 0 else 0
        scores[algo_key] = min(reduction_score + speed_score, 100)

    return scores


# -------------------------------------------------------------------------
# Comparison charts — timestamped, no overwrite (BUG-1 fix)
# -------------------------------------------------------------------------


def generate_reduction_chart(per_algorithm, output_dir):
    """Generate a bar chart for average reduction (%) with timestamp.

    Returns the path to the saved PNG file.
    """
    charts_dir = Path(output_dir)
    charts_dir.mkdir(parents=True, exist_ok=True)

    timestamp = _timestamp_ms()
    output_path = charts_dir / f"reduction_chart_{timestamp}.png"

    labels = []
    values = []
    for algo_key in ["deflate", "zopfli", "oxipng"]:
        data = per_algorithm.get(algo_key)
        if data and data["metrics"]:
            s = data["summary"]
            labels.append(data["label"])
            values.append(s["avg_reduction"])

    fig, ax = plt.subplots(figsize=(7, 5))
    if not labels:
        ax.text(0.5, 0.5, "No comparison data available",
                ha="center", va="center", fontsize=14)
        ax.set_xlim(0, 1)
        ax.set_ylim(0, 1)
    else:
        colors = ["#0078d4", "#107c10", "#d83b01"]
        bars = ax.bar(labels, values, color=colors[:len(labels)], width=0.5)
        ax.set_title("Average Reduction (%)", fontsize=12, fontweight="bold")
        ax.set_ylabel("Reduction (%)")
        ax.tick_params(axis="x", rotation=15)
        for bar, val in zip(bars, values):
            ax.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + 0.3,
                    f"{val:.1f}%", ha="center", va="bottom", fontsize=9)

    plt.tight_layout()
    fig.savefig(str(output_path), dpi=100, bbox_inches="tight")
    plt.close(fig)
    return output_path


def generate_time_chart(per_algorithm, output_dir):
    """Generate a bar chart for average time (ms) with timestamp.

    Returns the path to the saved PNG file.
    """
    charts_dir = Path(output_dir)
    charts_dir.mkdir(parents=True, exist_ok=True)

    timestamp = _timestamp_ms()
    output_path = charts_dir / f"time_chart_{timestamp}.png"

    labels = []
    values = []
    for algo_key in ["deflate", "zopfli", "oxipng"]:
        data = per_algorithm.get(algo_key)
        if data and data["metrics"]:
            s = data["summary"]
            labels.append(data["label"])
            values.append(s["avg_time_ms"])

    fig, ax = plt.subplots(figsize=(7, 5))
    if not labels:
        ax.text(0.5, 0.5, "No comparison data available",
                ha="center", va="center", fontsize=14)
        ax.set_xlim(0, 1)
        ax.set_ylim(0, 1)
    else:
        colors = ["#0078d4", "#107c10", "#d83b01"]
        bars = ax.bar(labels, values, color=colors[:len(labels)], width=0.5)
        ax.set_title("Average Time (ms)", fontsize=12, fontweight="bold")
        ax.set_ylabel("Time (ms)")
        ax.tick_params(axis="x", rotation=15)
        for bar, val in zip(bars, values):
            ax.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + 0.3,
                    f"{val:.1f}", ha="center", va="bottom", fontsize=9)

    plt.tight_layout()
    fig.savefig(str(output_path), dpi=100, bbox_inches="tight")
    plt.close(fig)
    return output_path


def _find_best_worst(metrics):
    """Find the best and worst file from a metrics list."""
    successful = [m for m in metrics if m["status"] == "SUCCESS"]
    if not successful:
        return "-", "-"

    best = max(successful, key=lambda m: m["reduction_percent"])
    worst = min(successful, key=lambda m: m["reduction_percent"])
    return best["file"], worst["file"]


def export_to_csv_new(metrics, summary, selected_algos, file_path):
    """Export actual running metrics and metadata to a specific CSV file."""
    output_path = Path(file_path)
    
    # Metadata calculations
    total_files = len(set(m["file"] for m in metrics))
    successful_files = len(set(m["file"] for m in metrics if m["status"] == "SUCCESS"))
    failed_files = total_files - successful_files
    
    selected_algos_str = "; ".join(selected_algos)
    
    with output_path.open("w", newline="", encoding="utf-8") as csv_file:
        writer = csv.writer(csv_file)
        
        # Write metadata
        writer.writerow(["Export Timestamp", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
        writer.writerow(["Application Name", "PNG Compression Comparison System"])
        writer.writerow(["Selected Algorithms", selected_algos_str])
        writer.writerow(["Total Files", total_files])
        writer.writerow(["Successful Files", successful_files])
        writer.writerow(["Failed Files", failed_files])
        writer.writerow(["Average Reduction", f"{summary.get('avg_reduction', 0.0):.2f}%"])
        avg_time_ms = summary.get("avg_time_ms", 0.0)
        if avg_time_ms >= 1000:
            avg_time_str = f"{avg_time_ms / 1000.0:.2f} s"
        else:
            avg_time_str = f"{avg_time_ms:.2f} ms"
        writer.writerow(["Average Processing Time", avg_time_str])
        
        writer.writerow([]) # blank line
        
        headers = [
            "File Name",
            "Original Size",
            "Algorithm",
            "Compressed Size",
            "Reduction %",
            "Compression Ratio",
            "Time (ms)",
            "Throughput"
        ]
        writer.writerow(headers)
        
        def format_size(bytes_val):
            if bytes_val <= 0:
                return "-"
            if bytes_val < 1024 * 1024:
                return f"{bytes_val / 1024:.2f} KB"
            return f"{bytes_val / (1024 * 1024):.2f} MB"
            
        for m in metrics:
            orig_sz = m["original_size"]
            comp_sz = m["compressed_size"]
            reduction = m["reduction_percent"]
            time_ms = m["time_ms"]
            algo = m["algorithm"]
            
            # Ratio
            ratio = orig_sz / comp_sz if comp_sz > 0 else 1.0
            
            # Throughput
            time_s = time_ms / 1000.0
            throughput_val = orig_sz / time_s if time_s > 0 else 0
            if throughput_val <= 0:
                tp_str = "-"
            elif throughput_val < 1024 * 1024:
                tp_str = f"{throughput_val / 1024:.1f} KB/s"
            else:
                tp_str = f"{throughput_val / (1024*1024):.1f} MB/s"
                
            writer.writerow([
                m["file"],
                format_size(orig_sz),
                algo,
                format_size(comp_sz),
                f"{reduction:.2f}%",
                f"{ratio:.2f}x",
                f"{time_ms:.2f}",
                tp_str
            ])
            
    return output_path


def export_to_excel(metrics, summary, per_algorithm, selected_algos, file_path):
    """Export metrics, summaries, rankings, and details to a multi-sheet XLSX file."""
    import openpyxl
    
    # Metadata calculations
    total_files = len(set(m["file"] for m in metrics))
    successful_files = len(set(m["file"] for m in metrics if m["status"] == "SUCCESS"))
    failed_files = total_files - successful_files
    selected_algos_str = "; ".join(selected_algos)
    
    wb = openpyxl.Workbook()
    
    # ----------------- SHEET 1: Metrics Table -----------------
    ws1 = wb.active
    ws1.title = "Metrics Table"
    
    metadata = [
        ("Export Timestamp", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Application Name", "PNG Compression Comparison System"),
        ("Selected Algorithms", selected_algos_str),
        ("Total Files", total_files),
        ("Successful Files", successful_files),
        ("Failed Files", failed_files),
        ("Average Reduction", f"{summary.get('avg_reduction', 0.0):.2f}%"),
    ]
    avg_time_ms = summary.get("avg_time_ms", 0.0)
    if avg_time_ms >= 1000:
        avg_time_str = f"{avg_time_ms / 1000.0:.2f} s"
    else:
        avg_time_str = f"{avg_time_ms:.2f} ms"
    metadata.append(("Average Processing Time", avg_time_str))
    
    for idx, (k, v) in enumerate(metadata, start=1):
        ws1.cell(row=idx, column=1, value=k)
        ws1.cell(row=idx, column=2, value=v)
        
    start_row = len(metadata) + 2
    
    headers1 = [
        "File Name",
        "Original Size",
        "Algorithm",
        "Compressed Size",
        "Reduction %",
        "Compression Ratio",
        "Time (ms)",
        "Throughput"
    ]
    for col_idx, h in enumerate(headers1, start=1):
        ws1.cell(row=start_row, column=col_idx, value=h)
        
    def format_size(bytes_val):
        if bytes_val <= 0:
            return "-"
        if bytes_val < 1024 * 1024:
            return f"{bytes_val / 1024:.2f} KB"
        return f"{bytes_val / (1024 * 1024):.2f} MB"
        
    row_idx = start_row + 1
    for m in metrics:
        orig_sz = m["original_size"]
        comp_sz = m["compressed_size"]
        reduction = m["reduction_percent"]
        time_ms = m["time_ms"]
        algo = m["algorithm"]
        
        # Ratio
        ratio = orig_sz / comp_sz if comp_sz > 0 else 1.0
        
        # Throughput
        time_s = time_ms / 1000.0
        throughput_val = orig_sz / time_s if time_s > 0 else 0
        if throughput_val <= 0:
            tp_str = "-"
        elif throughput_val < 1024 * 1024:
            tp_str = f"{throughput_val / 1024:.1f} KB/s"
        else:
            tp_str = f"{throughput_val / (1024*1024):.1f} MB/s"
            
        row_vals = [
            m["file"],
            format_size(orig_sz),
            algo,
            format_size(comp_sz),
            f"{reduction:.2f}%",
            f"{ratio:.2f}x",
            f"{time_ms:.2f}",
            tp_str
        ]
        for col_idx, val in enumerate(row_vals, start=1):
            ws1.cell(row=row_idx, column=col_idx, value=val)
        row_idx += 1

    # ----------------- SHEET 2: Comparison Summary -----------------
    ws2 = wb.create_sheet(title="Comparison Summary")
    
    success_files = [m for m in metrics if m["status"] == "SUCCESS"]
    avg_red_val = "-"
    best_red_val = "-"
    fastest_algo_val = "-"
    slowest_algo_val = "-"
    avg_time_val = "-"
    total_processed_val = "-"
    
    if success_files:
        reductions = [m["reduction_percent"] for m in success_files]
        times = [m["time_ms"] for m in metrics]
        
        avg_red_val = f"{sum(reductions) / len(reductions):.2f}%"
        best_red_val = f"{max(reductions):.2f}%"
        avg_time_val = f"{sum(times) / len(times):.2f} ms"
        total_processed_val = str(len(metrics))
        
        algo_times = {}
        for m in metrics:
            if m["status"] == "SUCCESS":
                algo = m["algorithm"]
                if algo not in algo_times:
                    algo_times[algo] = []
                algo_times[algo].append(m["time_ms"])
                
        if algo_times:
            avg_algo_times = {algo: sum(vals)/len(vals) for algo, vals in algo_times.items()}
            fastest_algo_val = min(avg_algo_times, key=avg_algo_times.get)
            slowest_algo_val = max(avg_algo_times, key=avg_algo_times.get)
            
    summary_data = [
        ("Average Reduction", avg_red_val),
        ("Best Reduction", best_red_val),
        ("Fastest Algorithm", fastest_algo_val),
        ("Slowest Algorithm", slowest_algo_val),
        ("Average Processing Time", avg_time_val),
        ("Total Files Processed", total_processed_val)
    ]
    
    ws2.cell(row=1, column=1, value="Metric Card")
    ws2.cell(row=1, column=2, value="Value")
    for idx, (k, v) in enumerate(summary_data, start=2):
        ws2.cell(row=idx, column=1, value=k)
        ws2.cell(row=idx, column=2, value=v)

    # ----------------- SHEET 3: Ranking Table -----------------
    ws3 = wb.create_sheet(title="Ranking Table")
    headers3 = ["Rank", "Algorithm", "Compression Efficiency", "Processing Speed", "Overall Score"]
    for col_idx, h in enumerate(headers3, start=1):
        ws3.cell(row=1, column=col_idx, value=h)
        
    ranking_rows = []
    if per_algorithm:
        scores = compute_ranking_scores(per_algorithm)
        for algo, data in per_algorithm.items():
            s = data["summary"]
            score = scores.get(algo, 0)
            ranking_rows.append((data["label"], s["avg_reduction"], s["avg_time_ms"], score))
        ranking_rows.sort(key=lambda r: r[3], reverse=True)
        
    for idx, (label, avg_red, avg_time, score) in enumerate(ranking_rows, start=1):
        ws3.cell(row=idx+1, column=1, value=f"#{idx}")
        ws3.cell(row=idx+1, column=2, value=label)
        ws3.cell(row=idx+1, column=3, value=f"{avg_red:.2f}% reduction")
        ws3.cell(row=idx+1, column=4, value=f"{avg_time:.2f} ms")
        ws3.cell(row=idx+1, column=5, value=f"{score:.2f} / 100")

    # ----------------- SHEET 4: File Details -----------------
    ws4 = wb.create_sheet(title="File Details")
    headers4 = [
        "File Name",
        "Original Size",
        "Compressed Size",
        "Reduction %",
        "Algorithm",
        "Processing Time",
        "Compression Ratio",
        "Status"
    ]
    for col_idx, h in enumerate(headers4, start=1):
        ws4.cell(row=1, column=col_idx, value=h)
        
    for idx, m in enumerate(metrics, start=2):
        orig_sz = m["original_size"]
        comp_sz = m["compressed_size"]
        reduction = m["reduction_percent"]
        time_ms = m["time_ms"]
        algo = m["algorithm"]
        status = m["status"]
        
        # Ratio
        ratio = orig_sz / comp_sz if comp_sz > 0 else 1.0
        
        ws4.cell(row=idx, column=1, value=m["file"])
        ws4.cell(row=idx, column=2, value=format_size(orig_sz))
        ws4.cell(row=idx, column=3, value=format_size(comp_sz))
        ws4.cell(row=idx, column=4, value=f"{reduction:.2f}%")
        ws4.cell(row=idx, column=5, value=algo)
        ws4.cell(row=idx, column=6, value=f"{time_ms:.2f} ms")
        ws4.cell(row=idx, column=7, value=f"{ratio:.2f}x")
        ws4.cell(row=idx, column=8, value=status)

    wb.save(file_path)
    return Path(file_path)


def _get_algorithm_label(algo):
    """Local helper to format algorithm keys into user-friendly names."""
    labels = {
        "deflate": "Deflate Baseline",
        "deflate_baseline": "Deflate Baseline",
        "zopfli": "Zopfli Deflate",
        "zopfli_deflate": "Zopfli Deflate",
        "oxipng": "OxiPNG",
    }
    return labels.get(str(algo).lower(), str(algo))


def export_to_csv_final(metrics, summary, selected_algos, file_path):
    """
    Export running metrics and metadata to a specific CSV file.
    Uses comma delimiter and UTF-8 encoding.
    """
    output_path = Path(file_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    
    export_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    formatted_algos = [_get_algorithm_label(a) for a in selected_algos]
    selected_algos_str = "; ".join(formatted_algos)
    
    unique_files = set(m["file"] for m in metrics)
    total_files_processed = len(unique_files)
    
    success_metrics = [m for m in metrics if m.get("status") == "SUCCESS"]
    if success_metrics:
        avg_reduction = sum(m["reduction_percent"] for m in success_metrics) / len(success_metrics)
    else:
        avg_reduction = 0.0
        
    if metrics:
        avg_time_ms = sum(m["time_ms"] for m in metrics) / len(metrics)
    else:
        avg_time_ms = 0.0
        
    with output_path.open("w", newline="", encoding="utf-8") as csv_file:
        writer = csv.writer(csv_file)
        
        # Write metadata
        writer.writerow(["Export Date", export_date])
        writer.writerow(["Selected Algorithms", selected_algos_str])
        writer.writerow(["Total Files Processed", total_files_processed])
        writer.writerow(["Average Reduction", f"{avg_reduction:.2f}%"])
        writer.writerow(["Average Processing Time", f"{avg_time_ms:.2f} ms"])
        
        # Blank row
        writer.writerow([])
        
        # Table Headers
        headers = [
            "File Name",
            "Algorithm",
            "Original Size (KB)",
            "Compressed Size (KB)",
            "Reduction (%)",
            "Compression Ratio",
            "Throughput (KB/s)",
            "Time (ms)"
        ]
        writer.writerow(headers)
        
        for m in metrics:
            fname = m["file"]
            algo = m["algorithm"]
            orig_sz = m["original_size"]
            comp_sz = m["compressed_size"]
            reduction = m["reduction_percent"]
            time_ms = m["time_ms"]
            status = m.get("status", "SUCCESS")
            
            orig_sz_kb = orig_sz / 1024.0
            comp_sz_kb = comp_sz / 1024.0
            
            if status == "SUCCESS":
                ratio = orig_sz / comp_sz if comp_sz > 0 else 1.0
                time_s = time_ms / 1000.0
                throughput = orig_sz_kb / time_s if time_s > 0 else 0.0
            else:
                ratio = 1.0
                throughput = 0.0
                comp_sz_kb = 0.0
                reduction = 0.0
                
            writer.writerow([
                fname,
                algo,
                f"{orig_sz_kb:.2f}",
                f"{comp_sz_kb:.2f}",
                f"{reduction:.2f}",
                f"{ratio:.2f}",
                f"{throughput:.2f}",
                f"{time_ms:.2f}"
            ])
            
    return output_path


def export_to_excel_final(metrics, summary, selected_algos, file_path):
    """
    Export metrics and summaries to a premium styled Excel (.xlsx) file with two sheets:
    - Detailed Results (equivalent to CSV Detailed Results)
    - Algorithm Summary (dynamic aggregation per algorithm)
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    
    # Premium Header style: Dark slate blue fill, white bold font
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Standard styling
    data_font = Font(name="Segoe UI", size=10)
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    align_center = Alignment(horizontal="center", vertical="center")
    
    meta_font_bold = Font(name="Segoe UI", size=10, bold=True)
    meta_font_val = Font(name="Segoe UI", size=10)
    
    thin_border = Border(
        left=Side(style='thin', color='E0E0E0'),
        right=Side(style='thin', color='E0E0E0'),
        top=Side(style='thin', color='E0E0E0'),
        bottom=Side(style='thin', color='E0E0E0')
    )

    # ----------------- SHEET 1: Detailed Results -----------------
    ws1 = wb.active
    ws1.title = "Detailed Results"
    ws1.views.sheetView[0].showGridLines = True
    
    # Calculate metadata values
    export_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    formatted_algos = [_get_algorithm_label(a) for a in selected_algos]
    selected_algos_str = "; ".join(formatted_algos)
    
    unique_files = set(m["file"] for m in metrics)
    total_files_processed = len(unique_files)
    
    success_metrics = [m for m in metrics if m.get("status") == "SUCCESS"]
    if success_metrics:
        avg_reduction = sum(m["reduction_percent"] for m in success_metrics) / len(success_metrics)
    else:
        avg_reduction = 0.0
        
    if metrics:
        avg_time_ms = sum(m["time_ms"] for m in metrics) / len(metrics)
    else:
        avg_time_ms = 0.0

    metadata = [
        ("Export Date", export_date),
        ("Selected Algorithms", selected_algos_str),
        ("Total Files Processed", total_files_processed),
        ("Average Reduction", f"{avg_reduction:.2f}%"),
        ("Average Processing Time", f"{avg_time_ms:.2f} ms"),
    ]
    
    for idx, (k, v) in enumerate(metadata, start=1):
        cell_k = ws1.cell(row=idx, column=1, value=k)
        cell_v = ws1.cell(row=idx, column=2, value=v)
        cell_k.font = meta_font_bold
        cell_v.font = meta_font_val
        cell_k.alignment = align_left
        cell_v.alignment = align_left

    start_row = len(metadata) + 2
    
    headers = [
        "File Name",
        "Algorithm",
        "Original Size (KB)",
        "Compressed Size (KB)",
        "Reduction (%)",
        "Compression Ratio",
        "Throughput (KB/s)",
        "Time (ms)"
    ]
    
    for col_idx, h in enumerate(headers, start=1):
        cell = ws1.cell(row=start_row, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = thin_border
        
    row_idx = start_row + 1
    for m in metrics:
        fname = m["file"]
        algo = m["algorithm"]
        orig_sz = m["original_size"]
        comp_sz = m["compressed_size"]
        reduction = m["reduction_percent"]
        time_ms = m["time_ms"]
        status = m.get("status", "SUCCESS")
        
        orig_sz_kb = orig_sz / 1024.0
        comp_sz_kb = comp_sz / 1024.0
        
        if status == "SUCCESS":
            ratio = orig_sz / comp_sz if comp_sz > 0 else 1.0
            time_s = time_ms / 1000.0
            throughput = orig_sz_kb / time_s if time_s > 0 else 0.0
        else:
            ratio = 1.0
            throughput = 0.0
            comp_sz_kb = 0.0
            reduction = 0.0
            
        row_vals = [
            fname,
            algo,
            round(orig_sz_kb, 2),
            round(comp_sz_kb, 2),
            round(reduction, 2),
            round(ratio, 2),
            round(throughput, 2),
            round(time_ms, 2)
        ]
        
        for col_idx, val in enumerate(row_vals, start=1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=val)
            cell.font = data_font
            cell.border = thin_border
            
            if col_idx in [1, 2]:
                cell.alignment = align_left
            else:
                cell.alignment = align_right
                cell.number_format = '0.00'
                
        row_idx += 1

    # Auto-adjust column widths for Detailed Results sheet
    for col in ws1.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row < start_row:
                continue
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws1.column_dimensions[col_letter].width = max(max_len + 4, 15)
        
    ws1.column_dimensions['A'].width = 25
    ws1.column_dimensions['B'].width = 25

    # ----------------- SHEET 2: Algorithm Summary -----------------
    ws2 = wb.create_sheet(title="Algorithm Summary")
    ws2.views.sheetView[0].showGridLines = True
    
    headers_summary = [
        "Algorithm",
        "Files Processed",
        "Average Reduction (%)",
        "Average Compression Ratio",
        "Average Throughput (KB/s)",
        "Average Processing Time (ms)"
    ]
    
    for col_idx, h in enumerate(headers_summary, start=1):
        cell = ws2.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = thin_border
        
    # Group by algorithm key present in metrics
    algo_groups = {}
    for m in metrics:
        algo_name = m["algorithm"]
        if algo_name not in algo_groups:
            algo_groups[algo_name] = []
        algo_groups[algo_name].append(m)
        
    row_idx = 2
    for algo_name, algo_metrics in sorted(algo_groups.items()):
        total_processed = len(algo_metrics)
        success_algo = [m for m in algo_metrics if m.get("status") == "SUCCESS"]
        
        # Avg reduction (%)
        if success_algo:
            avg_red = sum(m["reduction_percent"] for m in success_algo) / len(success_algo)
        else:
            avg_red = 0.0
            
        # Avg compression ratio
        ratios = []
        for m in success_algo:
            orig = m["original_size"]
            comp = m["compressed_size"]
            ratios.append(orig / comp if comp > 0 else 1.0)
        avg_ratio = sum(ratios) / len(ratios) if ratios else 1.0
        
        # Avg throughput (KB/s)
        throughputs = []
        for m in success_algo:
            orig_sz_kb = m["original_size"] / 1024.0
            time_s = m["time_ms"] / 1000.0
            if time_s > 0:
                throughputs.append(orig_sz_kb / time_s)
        avg_throughput = sum(throughputs) / len(throughputs) if throughputs else 0.0
        
        # Avg processing time (ms)
        avg_time = sum(m["time_ms"] for m in algo_metrics) / len(algo_metrics) if algo_metrics else 0.0
        
        row_vals = [
            algo_name,
            total_processed,
            round(avg_red, 2),
            round(avg_ratio, 2),
            round(avg_throughput, 2),
            round(avg_time, 2)
        ]
        
        for col_idx, val in enumerate(row_vals, start=1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=val)
            cell.font = data_font
            cell.border = thin_border
            
            if col_idx == 1:
                cell.alignment = align_left
            elif col_idx == 2:
                cell.alignment = align_center
                cell.number_format = '#,##0'
            else:
                cell.alignment = align_right
                cell.number_format = '0.00'
                
        row_idx += 1

    # Auto-adjust column widths for Algorithm Summary sheet
    for col in ws2.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws2.column_dimensions[col_letter].width = max(max_len + 4, 18)

    wb.save(file_path)
    return Path(file_path)

