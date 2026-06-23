"""
Unit tests for src.export.exporter
"""

import unittest
import unittest.mock
import tempfile
import shutil
from pathlib import Path

from src.export.exporter import (
    export_metrics_csv,
    _build_report_path,
    _format_metric_row,
    _bytes_to_kb,
    CSV_COLUMNS,
)


class TestExportMetricsCsv(unittest.TestCase):
    """Tests for export_metrics_csv()"""

    def setUp(self):
        self.tmpdir = Path(tempfile.mkdtemp())
        self.metrics = [
            {
                "file": "a.png",
                "algorithm": "Deflate Baseline",
                "original_size": 20480,
                "compressed_size": 10240,
                "reduction_percent": 50.0,
                "time_ms": 150.5,
                "resolution": "16x16",
                "status": "SUCCESS",
            },
            {
                "file": "b.png",
                "algorithm": "Deflate Baseline",
                "original_size": 102400,
                "compressed_size": 51200,
                "reduction_percent": 50.0,
                "time_ms": 200.0,
                "resolution": "32x32",
                "status": "SUCCESS",
            },
        ]
        self.summary = {
            "completed": 2,
            "failed": 0,
            "cancelled": False,
            "avg_reduction": 50.0,
            "avg_time_ms": 175.25,
        }

    def tearDown(self):
        shutil.rmtree(self.tmpdir, ignore_errors=True)

    # ---------- happy path ----------

    def test_export_creates_csv(self):
        path = export_metrics_csv(self.metrics, self.summary, self.tmpdir)
        self.assertTrue(path.is_file())
        self.assertEqual(path.name, "results_export.csv")
        self.assertEqual(path.suffix, ".csv")

    def test_export_contains_header(self):
        path = export_metrics_csv(self.metrics, self.summary, self.tmpdir)
        content = path.read_text(encoding="utf-8")
        headers = [
            "File Name",
            "Original Size",
            "Algorithm",
            "Compressed Size",
            "Reduction %",
            "Compression Ratio",
            "Processing Time (ms)",
            "Best Algorithm",
            "Score",
            "Timestamp"
        ]
        for col in headers:
            self.assertIn(col, content)

    def test_export_contains_data_rows(self):
        path = export_metrics_csv(self.metrics, self.summary, self.tmpdir)
        content = path.read_text(encoding="utf-8")
        self.assertIn("a.png", content)
        self.assertIn("b.png", content)

    # ---------- empty metrics ----------

    def test_export_empty_metrics(self):
        path = export_metrics_csv([], self.summary, self.tmpdir)
        self.assertTrue(path.is_file())
        content = path.read_text(encoding="utf-8")
        self.assertIn("File Name,Original Size,Algorithm", content)

    # ---------- output directory ----------

    def test_export_creates_output_dir(self):
        new_dir = self.tmpdir / "deep" / "nested" / "reports"
        self.assertFalse(new_dir.exists())
        path = export_metrics_csv(self.metrics, self.summary, new_dir)
        self.assertTrue(path.is_file())
        self.assertTrue(new_dir.is_dir())


class TestBuildReportPath(unittest.TestCase):
    """Tests for _build_report_path()"""

    def setUp(self):
        self.tmpdir = Path(tempfile.mkdtemp())

    def tearDown(self):
        shutil.rmtree(self.tmpdir, ignore_errors=True)

    def test_returns_timestamped(self):
        path = _build_report_path(self.tmpdir)
        self.assertIn("report_", path.name)
        self.assertEqual(path.suffix, ".csv")

    def test_appends_counter_on_collision(self):
        fixed_ts = "20260617_163122_483"
        with unittest.mock.patch(
            "src.export.exporter._timestamp_ms", return_value=fixed_ts
        ):
            p1 = _build_report_path(self.tmpdir)
            p1.touch()  # occupy first name
            p2 = _build_report_path(self.tmpdir)
        self.assertNotEqual(p1, p2)
        self.assertIn("_1.", p2.name)


class TestFormatMetricRow(unittest.TestCase):
    """Tests for _format_metric_row()"""

    def test_format_converts_bytes_to_kb(self):
        metric = {
            "file": "a.png",
            "algorithm": "Deflate",
            "original_size": 20480,
            "compressed_size": 10240,
            "reduction_percent": 50.0,
            "time_ms": 100.0,
            "resolution": "10x10",
            "status": "SUCCESS",
        }
        row = _format_metric_row(metric)
        self.assertEqual(row["file"], "a.png")
        self.assertIn(".", row["original_size_kb"])  # float format
        self.assertNotIn("KB", row["compressed_size_kb"])


class TestBytesToKb(unittest.TestCase):
    """Tests for _bytes_to_kb()"""

    def test_positive_value(self):
        self.assertEqual(_bytes_to_kb(1024), "1.0000")

    def test_zero(self):
        self.assertEqual(_bytes_to_kb(0), "0.0000")

    def test_negative(self):
        self.assertEqual(_bytes_to_kb(-100), "0.0000")

    def test_large_value(self):
        result = _bytes_to_kb(1048576)
        self.assertEqual(result, "1024.0000")


from src.export.exporter import export_to_csv_final, export_to_excel_final

class TestNewExporters(unittest.TestCase):
    def setUp(self):
        self.tmpdir = Path(tempfile.mkdtemp())
        self.metrics = [
            {
                "file": "a.png",
                "algorithm": "Deflate Baseline",
                "original_size": 20480,
                "compressed_size": 10240,
                "reduction_percent": 50.0,
                "time_ms": 150.5,
                "resolution": "16x16",
                "status": "SUCCESS",
            },
            {
                "file": "b.png",
                "algorithm": "Zopfli Deflate",
                "original_size": 102400,
                "compressed_size": 51200,
                "reduction_percent": 50.0,
                "time_ms": 200.0,
                "resolution": "32x32",
                "status": "SUCCESS",
            },
        ]
        self.summary = {
            "completed": 2,
            "failed": 0,
            "cancelled": False,
            "avg_reduction": 50.0,
            "avg_time_ms": 175.25,
        }
        self.per_algorithm = {
            "deflate": {
                "label": "Deflate Baseline",
                "summary": {"completed": 1, "failed": 0, "avg_reduction": 50.0, "avg_time_ms": 150.5},
                "metrics": [self.metrics[0]]
            },
            "zopfli": {
                "label": "Zopfli Deflate",
                "summary": {"completed": 1, "failed": 0, "avg_reduction": 50.0, "avg_time_ms": 200.0},
                "metrics": [self.metrics[1]]
            }
        }
        self.selected_algos = ["deflate", "zopfli"]

    def tearDown(self):
        shutil.rmtree(self.tmpdir, ignore_errors=True)

    def test_export_to_csv_final(self):
        dest = self.tmpdir / "results.csv"
        path = export_to_csv_final(self.metrics, self.summary, self.selected_algos, dest)
        self.assertTrue(path.is_file())
        content = path.read_text(encoding="utf-8")
        self.assertIn("Export Date", content)
        self.assertIn("a.png", content)
        self.assertIn("b.png", content)

    def test_export_to_excel_final(self):
        dest = self.tmpdir / "results.xlsx"
        path = export_to_excel_final(self.metrics, self.summary, self.selected_algos, dest)
        self.assertTrue(path.is_file())
        import openpyxl
        wb = openpyxl.load_workbook(dest)
        self.assertIn("Detailed Results", wb.sheetnames)
        self.assertIn("Algorithm Summary", wb.sheetnames)


if __name__ == "__main__":
    unittest.main()
